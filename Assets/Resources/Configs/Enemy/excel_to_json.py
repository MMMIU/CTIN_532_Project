import openpyxl
import json

# excel to json
def excel_to_json(excel_file, json_f_name):
    jd = []
    heads = []
    book = openpyxl.load_workbook(excel_file)
    sheet = book[u'Sheet1']
    
    max_row = sheet.max_row
    max_column = sheet.max_column
    # analyze headers
    for column in range(max_column):
        heads.append(sheet.cell(1, column + 1).value)
    # for each row
    for row in range(max_row):
        if row < 1:
        	# skip header line
            continue
        one_line = {}
        # for each column
        for column in range(max_column): 
            k = heads[column]
            v = sheet.cell(row + 1, column + 1).value
            one_line[k] = v
        jd.append(one_line)
    book.close()
    save_json_file(jd, json_f_name)

# save json to file
def save_json_file(jd, json_f_name):
    f = open(json_f_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    f.write(txt)
    f.close()

if '__main__' == __name__:
     excel_to_json(u'Enemies.xlsx', 'enemy_cfg.bytes')
